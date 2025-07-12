import os
import torch
import argparse
import openpyxl

from data import load_data
from embedding import get_embeddings, get_logits
from model import load_model, load_decoder_layer
from utils import check_args

def calculate_metric(args, data, tokenizer, topk_indices):
    
    # def filter_token_list(token_list):
    #     for w in [101, 102]:
    #         if w in token_list:
    #             token_list.remove(w)
    #     return token_list
    
    avg_hit_k, avg_local_align = 0, 0
    valid_num = 0
    total_decode_token, total_match_token, total_exist_token = set(), set(), set()
    for i, sent in enumerate(data):
        token_list = tokenizer.encode(sent)
        # token_list = filter_token_list(token_list)
        topk_idxs = topk_indices[i]
        if len(token_list) == 0:
            continue
        valid_num += 1
        hit_k = int(len(set(topk_idxs[:args.k]) & set(token_list)) > 0)
        local_align = len(set(topk_idxs[:len(token_list)]) & set(token_list)) / len(set(token_list))
        total_decode_token = total_decode_token | set(topk_idxs)
        total_match_token = total_match_token | (set(topk_idxs[:len(token_list)]) & set(token_list))
        total_exist_token = total_exist_token | set(token_list)
        avg_hit_k += hit_k
        avg_local_align += local_align
    avg_hit_k /= valid_num
    avg_local_align /= valid_num
    avg_global_align = len(total_match_token) / len(total_exist_token)

    return avg_hit_k, avg_local_align, avg_global_align


def get_decode_content(tokenizer, logits, k):
    topk_tokens = []
    _, topk_indices = torch.topk(logits, k)
    topk_indices = topk_indices.numpy().tolist()
    for i in range(len(topk_indices)):
        each_dot_max_tokens = [tokenizer.convert_ids_to_tokens(idx) for idx in topk_indices[i]]
        topk_tokens.append(each_dot_max_tokens)
    return topk_indices, topk_tokens


def obtain_aligned_tokens(args, texts, model_list):

    model_name, data_name = args.model, args.dataset
    
    if data_name == "msmarco":
        half_size = int(len(texts) / 2)
        texts_q, texts_d = texts[:half_size], texts[half_size:]
        embs_q = get_embeddings(model_name=model_name, model_list=model_list, data_name=data_name, texts=texts_q).cuda()
        embs_d = get_embeddings(model_name=model_name, model_list=model_list, data_name=data_name, texts=texts_d, isdoc=True).cuda()
        embs = torch.cat((embs_q, embs_d), dim=0)
    else:
        embs = get_embeddings(model_name=model_name, model_list=model_list, data_name=data_name, texts=texts).cuda()
    
    decoder_layer = load_decoder_layer(args, model=model)
    logits = get_logits(args, decoder_layer, embs)
    topk_indices, topk_tokens = get_decode_content(tokenizer=tokenizer, logits=logits, k=args.k)

    return texts, topk_tokens, topk_indices


def save_xlsx(args, save_path, info_list, metric_list):
    texts, decoded_tokens = info_list
    if args.dataset=="wiki":
        printed_list = info_list
    else:
        half_size = int(len(texts) / 2)
        printed_list = [texts[:half_size], texts[half_size:], decoded_tokens[:half_size], decoded_tokens[half_size:]]
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    
    
    for j in range(len(metric_list)):
        sheet.cell(row=1, column=j + 1, value=str(metric_list[j]))
    
    for index, tuple in enumerate(zip(*printed_list)):
        for j in range(len(printed_list)):
            sheet.cell(row=index + 2, column=j + 1, value=str(tuple[j]))
    workbook.save(save_path)


if __name__ == "__main__":

    # "bert", "prompt_bert", "simcse", "contriever", "dpr", "sgpt_nli", "sgpt_msmarco",
    # "opt_eol", "gte", "e5", "llama_eol_cse", "opt_eol_cse", "e5_mistral", "syncse",
    # "mistral", "llm2vec_mistral", "GritLM"
    
    parse = argparse.ArgumentParser()
    parse.add_argument('--model', type=str, help='the embedder name', default="sgpt_nli")
    parse.add_argument('--dataset', type=str, help='the dataset name', choices=["sts", "msmarco", "nli", "wiki"], default="sts")
    parse.add_argument('--llama_path', type=str, help='the name of dataset', default="/data/LLM/llama1/llama-7b")
    parse.add_argument('--weight_dir', type=str, help='the dir to save the lm_head weight', default="./save/model_weights/")
    parse.add_argument('--output_dir', type=str, help='the dir to save the xlsx file', default="./output/aligned_tokens/")
    parse.add_argument('--k', type=int, help='the number of aligned token', default=10)
    
    args = parse.parse_args()
    
    check_args(args)
    
    save_path = f"{args.model}_{args.dataset}_results.xlsx"
    if os.path.exists(save_path):
        print(f"Existing file {save_path} in the directory {args.output_dir}")
        exit(0)
    else:
        args.save_path = save_path
        
        
    texts = load_data(data_name=args.dataset)
    model, tokenizer, doc_model, doc_tokenizer = load_model(args, model_name=args.model)
    model_list = [model, tokenizer, doc_model, doc_tokenizer]
    
    texts, topk_tokens, topk_indices = obtain_aligned_tokens(args, texts, model_list)
    avg_hit_k, avg_local_align, avg_global_align = calculate_metric(args, texts, tokenizer, topk_indices)
    print(avg_hit_k, avg_local_align, avg_global_align)
    save_xlsx(args, os.path.join(args.output_dir, args.save_path), info_list=[texts, topk_tokens], metric_list=[avg_hit_k, avg_local_align, avg_global_align])
    
    